"""Admin activity logs and monitoring endpoints"""
from fastapi import APIRouter, Depends, HTTPException, Query, Response, WebSocket, WebSocketDisconnect
from sqlmodel import Session, select, func, or_, and_
from typing import Optional, List
from datetime import datetime, timedelta
from database import get_session
from models import User, ActivityLog
from dependencies import get_current_user, require_admin
from pydantic import BaseModel
import csv
import io
import json
import asyncio

router = APIRouter(prefix="/api/admin/activity-logs", tags=["Admin Activity Logs"])

# WebSocket connection manager for real-time logs
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                pass

manager = ConnectionManager()

# Request/Response Models
class ActivityLogFilter(BaseModel):
    user_name: Optional[str] = None
    user_id: Optional[int] = None
    role: Optional[str] = None
    activity_type: Optional[str] = None
    date_from: Optional[str] = None  # ISO format date
    date_to: Optional[str] = None
    ip_address: Optional[str] = None
    device_type: Optional[str] = None
    status: Optional[str] = None
    keyword: Optional[str] = None

class ActivityLogResponse(BaseModel):
    id: int
    user_id: int
    user_name: str
    user_role: str
    activity_type: str
    activity_description: str
    ip_address: Optional[str]
    device_type: Optional[str]
    timestamp: str

@router.get("/")
async def get_activity_logs(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    user_name: Optional[str] = None,
    user_id: Optional[int] = None,
    role: Optional[str] = None,
    activity_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    ip_address: Optional[str] = None,
    device_type: Optional[str] = None,
    keyword: Optional[str] = None,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_session)
):
    """Get activity logs with advanced filtering (Admin only)"""
    statement = select(ActivityLog)
    
    # Apply filters
    conditions = []
    
    if user_name:
        conditions.append(ActivityLog.user_name.ilike(f"%{user_name}%"))
    
    if user_id:
        conditions.append(ActivityLog.user_id == user_id)
    
    if role:
        conditions.append(ActivityLog.user_role == role)
    
    if activity_type:
        conditions.append(ActivityLog.activity_type == activity_type)
    
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from)
            conditions.append(ActivityLog.timestamp >= date_from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format")
    
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to)
            conditions.append(ActivityLog.timestamp <= date_to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format")
    
    if ip_address:
        conditions.append(ActivityLog.ip_address.ilike(f"%{ip_address}%"))
    
    if device_type:
        conditions.append(ActivityLog.device_type == device_type)
    
    if keyword:
        conditions.append(
            or_(
                ActivityLog.activity_description.ilike(f"%{keyword}%"),
                ActivityLog.user_name.ilike(f"%{keyword}%"),
                ActivityLog.activity_type.ilike(f"%{keyword}%")
            )
        )
    
    if conditions:
        statement = statement.where(and_(*conditions))
    
    # Get total count
    count_statement = select(func.count()).select_from(ActivityLog)
    if conditions:
        count_statement = count_statement.where(and_(*conditions))
    total_count = db.exec(count_statement).one()
    
    # Get logs
    statement = statement.order_by(ActivityLog.timestamp.desc()).offset(skip).limit(limit)
    logs = db.exec(statement).all()
    
    return {
        "total": total_count,
        "skip": skip,
        "limit": limit,
        "logs": [
            {
                "id": log.id,
                "user_id": log.user_id,
                "user_name": log.user_name,
                "user_role": log.user_role,
                "activity_type": log.activity_type,
                "activity_description": log.activity_description,
                "ip_address": log.ip_address,
                "device_type": log.device_type,
                "user_agent": log.user_agent,
                "timestamp": log.timestamp.isoformat()
            }
            for log in logs
        ]
    }


@router.get("/stats")
async def get_activity_stats(
    days: int = Query(7, ge=1, le=90),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_session)
):
    """Get activity statistics for dashboard (Admin only)"""
    # Calculate date range
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)
    
    # Total activities
    total_statement = select(func.count()).select_from(ActivityLog).where(
        ActivityLog.timestamp >= start_date
    )
    total_activities = db.exec(total_statement).one()
    
    # Activities by type
    type_statement = select(
        ActivityLog.activity_type,
        func.count(ActivityLog.id).label("count")
    ).where(
        ActivityLog.timestamp >= start_date
    ).group_by(ActivityLog.activity_type)
    
    activities_by_type = []
    for row in db.exec(type_statement).all():
        activities_by_type.append({
            "activity_type": row[0],
            "count": row[1]
        })
    
    # Activities by role
    role_statement = select(
        ActivityLog.user_role,
        func.count(ActivityLog.id).label("count")
    ).where(
        ActivityLog.timestamp >= start_date
    ).group_by(ActivityLog.user_role)
    
    activities_by_role = []
    for row in db.exec(role_statement).all():
        activities_by_role.append({
            "role": row[0],
            "count": row[1]
        })
    
    # Activities by device
    device_statement = select(
        ActivityLog.device_type,
        func.count(ActivityLog.id).label("count")
    ).where(
        ActivityLog.timestamp >= start_date
    ).group_by(ActivityLog.device_type)
    
    activities_by_device = []
    for row in db.exec(device_statement).all():
        activities_by_device.append({
            "device_type": row[0],
            "count": row[1]
        })
    
    # Most active users
    user_statement = select(
        ActivityLog.user_id,
        ActivityLog.user_name,
        func.count(ActivityLog.id).label("count")
    ).where(
        ActivityLog.timestamp >= start_date
    ).group_by(
        ActivityLog.user_id,
        ActivityLog.user_name
    ).order_by(
        func.count(ActivityLog.id).desc()
    ).limit(10)
    
    most_active_users = []
    for row in db.exec(user_statement).all():
        most_active_users.append({
            "user_id": row[0],
            "user_name": row[1],
            "activity_count": row[2]
        })
    
    # Daily activity trend
    daily_statement = select(
        func.date(ActivityLog.timestamp).label("date"),
        func.count(ActivityLog.id).label("count")
    ).where(
        ActivityLog.timestamp >= start_date
    ).group_by(
        func.date(ActivityLog.timestamp)
    ).order_by(
        func.date(ActivityLog.timestamp)
    )
    
    daily_trend = []
    for row in db.exec(daily_statement).all():
        daily_trend.append({
            "date": row[0].isoformat() if row[0] else None,
            "count": row[1]
        })
    
    return {
        "period_days": days,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "total_activities": total_activities,
        "activities_by_type": activities_by_type,
        "activities_by_role": activities_by_role,
        "activities_by_device": activities_by_device,
        "most_active_users": most_active_users,
        "daily_trend": daily_trend
    }


@router.get("/export")
async def export_activity_logs(
    format: str = Query("csv", regex="^(csv|json)$"),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    role: Optional[str] = None,
    activity_type: Optional[str] = None,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_session)
):
    """Export activity logs to CSV or JSON (Admin only)"""
    statement = select(ActivityLog)
    
    # Apply filters
    conditions = []
    
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from)
            conditions.append(ActivityLog.timestamp >= date_from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format")
    
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to)
            conditions.append(ActivityLog.timestamp <= date_to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format")
    
    if role:
        conditions.append(ActivityLog.user_role == role)
    
    if activity_type:
        conditions.append(ActivityLog.activity_type == activity_type)
    
    if conditions:
        statement = statement.where(and_(*conditions))
    
    statement = statement.order_by(ActivityLog.timestamp.desc()).limit(10000)  # Limit to 10k records
    logs = db.exec(statement).all()
    
    if format == "csv":
        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "ID", "User ID", "User Name", "Role", "Activity Type",
            "Description", "IP Address", "Device Type", "Timestamp"
        ])
        
        # Write data
        for log in logs:
            writer.writerow([
                log.id,
                log.user_id,
                log.user_name,
                log.user_role,
                log.activity_type,
                log.activity_description,
                log.ip_address or "",
                log.device_type or "",
                log.timestamp.isoformat()
            ])
        
        csv_content = output.getvalue()
        output.close()
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=activity_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            }
        )
    
    else:  # JSON format
        logs_data = [
            {
                "id": log.id,
                "user_id": log.user_id,
                "user_name": log.user_name,
                "user_role": log.user_role,
                "activity_type": log.activity_type,
                "activity_description": log.activity_description,
                "ip_address": log.ip_address,
                "device_type": log.device_type,
                "timestamp": log.timestamp.isoformat()
            }
            for log in logs
        ]
        
        return {
            "export_date": datetime.utcnow().isoformat(),
            "total_records": len(logs_data),
            "logs": logs_data
        }


@router.get("/user/{user_id}")
async def get_user_activity_logs(
    user_id: int,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_session)
):
    """Get activity logs for a specific user (Admin only)"""
    # Check if user exists
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get logs
    statement = (
        select(ActivityLog)
        .where(ActivityLog.user_id == user_id)
        .order_by(ActivityLog.timestamp.desc())
        .offset(skip)
        .limit(limit)
    )
    logs = db.exec(statement).all()
    
    # Get total count
    count_statement = select(func.count()).select_from(ActivityLog).where(ActivityLog.user_id == user_id)
    total_count = db.exec(count_statement).one()
    
    return {
        "user": {
            "id": user.id,
            "name": user.full_name,
            "email": user.email,
            "role": user.role.value
        },
        "total": total_count,
        "skip": skip,
        "limit": limit,
        "logs": [
            {
                "id": log.id,
                "activity_type": log.activity_type,
                "activity_description": log.activity_description,
                "ip_address": log.ip_address,
                "device_type": log.device_type,
                "timestamp": log.timestamp.isoformat()
            }
            for log in logs
        ]
    }


@router.get("/export/excel")
async def export_activity_logs_excel(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    role: Optional[str] = None,
    activity_type: Optional[str] = None,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_session)
):
    """Export activity logs to Excel format (Admin only)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl library not installed")
    
    statement = select(ActivityLog)
    
    # Apply filters
    conditions = []
    
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from)
            conditions.append(ActivityLog.timestamp >= date_from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format")
    
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to)
            conditions.append(ActivityLog.timestamp <= date_to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format")
    
    if role:
        conditions.append(ActivityLog.user_role == role)
    
    if activity_type:
        conditions.append(ActivityLog.activity_type == activity_type)
    
    if conditions:
        statement = statement.where(and_(*conditions))
    
    statement = statement.order_by(ActivityLog.timestamp.desc()).limit(10000)
    logs = db.exec(statement).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activity Logs"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["ID", "User ID", "User Name", "Role", "Activity Type", 
               "Description", "IP Address", "Device Type", "User Agent", "Timestamp"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, log in enumerate(logs, 2):
        ws.cell(row=row_num, column=1, value=log.id).border = thin_border
        ws.cell(row=row_num, column=2, value=log.user_id).border = thin_border
        ws.cell(row=row_num, column=3, value=log.user_name).border = thin_border
        ws.cell(row=row_num, column=4, value=log.user_role).border = thin_border
        ws.cell(row=row_num, column=5, value=log.activity_type).border = thin_border
        ws.cell(row=row_num, column=6, value=log.activity_description).border = thin_border
        ws.cell(row=row_num, column=7, value=log.ip_address or "").border = thin_border
        ws.cell(row=row_num, column=8, value=log.device_type or "").border = thin_border
        ws.cell(row=row_num, column=9, value=log.user_agent or "").border = thin_border
        ws.cell(row=row_num, column=10, value=log.timestamp.isoformat()).border = thin_border
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Description column wider
    ws.column_dimensions['F'].width = 40
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=activity_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )


@router.get("/activity-types")
async def get_activity_types(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_session)
):
    """Get list of all activity types for filtering (Admin only)"""
    statement = select(ActivityLog.activity_type).distinct()
    types = db.exec(statement).all()
    return {"activity_types": [t for t in types if t]}


@router.get("/roles")
async def get_activity_roles(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_session)
):
    """Get list of all roles in activity logs (Admin only)"""
    statement = select(ActivityLog.user_role).distinct()
    roles = db.exec(statement).all()
    return {"roles": [r for r in roles if r]}


@router.get("/hourly-distribution")
async def get_hourly_distribution(
    days: int = Query(7, ge=1, le=30),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_session)
):
    """Get hourly activity distribution (Admin only)"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)
    
    statement = select(
        func.extract('hour', ActivityLog.timestamp).label("hour"),
        func.count(ActivityLog.id).label("count")
    ).where(
        ActivityLog.timestamp >= start_date
    ).group_by(
        func.extract('hour', ActivityLog.timestamp)
    ).order_by(
        func.extract('hour', ActivityLog.timestamp)
    )
    
    results = db.exec(statement).all()
    
    # Initialize all 24 hours with 0
    distribution = {i: 0 for i in range(24)}
    for row in results:
        hour = int(row[0]) if row[0] is not None else 0
        distribution[hour] = row[1]
    
    return {
        "period_days": days,
        "distribution": [{"hour": h, "count": c} for h, c in distribution.items()]
    }


@router.get("/search")
async def search_activity_logs(
    q: str = Query(..., min_length=2),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_session)
):
    """Full-text search across activity logs (Admin only)"""
    search_pattern = f"%{q}%"
    
    statement = select(ActivityLog).where(
        or_(
            ActivityLog.user_name.ilike(search_pattern),
            ActivityLog.activity_type.ilike(search_pattern),
            ActivityLog.activity_description.ilike(search_pattern),
            ActivityLog.ip_address.ilike(search_pattern)
        )
    ).order_by(ActivityLog.timestamp.desc()).offset(skip).limit(limit)
    
    logs = db.exec(statement).all()
    
    # Count total results
    count_statement = select(func.count()).select_from(ActivityLog).where(
        or_(
            ActivityLog.user_name.ilike(search_pattern),
            ActivityLog.activity_type.ilike(search_pattern),
            ActivityLog.activity_description.ilike(search_pattern),
            ActivityLog.ip_address.ilike(search_pattern)
        )
    )
    total = db.exec(count_statement).one()
    
    return {
        "query": q,
        "total": total,
        "logs": [
            {
                "id": log.id,
                "user_id": log.user_id,
                "user_name": log.user_name,
                "user_role": log.user_role,
                "activity_type": log.activity_type,
                "activity_description": log.activity_description,
                "ip_address": log.ip_address,
                "device_type": log.device_type,
                "timestamp": log.timestamp.isoformat()
            }
            for log in logs
        ]
    }


@router.websocket("/realtime")
async def websocket_realtime_logs(websocket: WebSocket):
    """WebSocket endpoint for real-time activity log streaming"""
    await manager.connect(websocket)
    try:
        while True:
            # Keep connection alive
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager.disconnect(websocket)


# Helper function to broadcast new activity logs
async def broadcast_activity_log(log_data: dict):
    """Broadcast a new activity log to all connected WebSocket clients"""
    await manager.broadcast({
        "type": "new_activity",
        "data": log_data
    })
